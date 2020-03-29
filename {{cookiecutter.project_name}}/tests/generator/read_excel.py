import openpyxl
from pick import pick
import os

def camel_to_underline(camel_format):
    '''
        驼峰命名格式转下划线命名格式
    '''
    underline_format = ''
    is_first=True
    if isinstance(camel_format, str):
        for _s_ in camel_format:
            if is_first:
                underline_format += _s_ if not _s_.isupper() else _s_.lower()
                is_first = False
            else:
                underline_format += _s_ if not _s_.isupper() else '_' + _s_.lower()
    return underline_format


def print_field_info_list(type: str,file_name: None, dbchange = False, use_udf = False):
    wb = openpyxl.load_workbook(file_name if not None else 'material.xlsx')  # 加载工作薄
    sheet = wb['Sheet1']  # 工作表

    db_field_list = []
    for i in range(1, 200):
        src_field_name: str = sheet.cell(row=i, column=2).value
        field_name = camel_to_underline(src_field_name)
        field_desc1: str = sheet.cell(row=i, column=1).value
        field_desc2: str = sheet.cell(row=i, column=4).value
        field_desc3: str = sheet.cell(row=i, column=6).value
        desc_set = set()
        desc = ''
        if (not desc_set.__contains__(field_desc1)) and field_desc1:
            desc_set.add(field_desc1.strip(' '))
            desc += ' ' + field_desc1.strip(' ')
        if (not desc_set.__contains__(field_desc2)) and field_desc2:
            desc_set.add(field_desc2.strip(' '))
            desc += ' ' + field_desc2.strip(' ')
        if (not desc_set.__contains__(field_desc3)) and field_desc3:
            desc_set.add(field_desc3.strip(' '))
            desc += ' ' + field_desc3.strip(' ')
        desc = desc.strip(' ')
        need = sheet.cell(row=i, column=5).value
        nullable: bool = True
        if need == '是':
            nullable = False
        else:
            nullable = True
        full_field_type: str = sheet.cell(row=i, column=3).value
        simple_field_type: str = full_field_type
        if not full_field_type:
            continue
        full_field_type = full_field_type.lower()
        str_count = 0
        if full_field_type.__contains__('varchar'):
            full_field_type = full_field_type.replace('varchar', 'String')
            simple_field_type = 'String'
        if full_field_type.__contains__('char'):
            full_field_type = full_field_type.replace('char', 'String')
            simple_field_type = 'String'
        if full_field_type.__contains__('int'):
            # full_field_type = re.sub('[(]\\d+[)]',  '', full_field_type, count=1)
            full_field_type = 'Integer()'
            simple_field_type = 'UdfInt' if use_udf and nullable else 'Integer'
        if full_field_type.__contains__('decimal'):
            full_field_type = full_field_type.replace('decimal', 'DECIMAL')
            simple_field_type = 'UdfDecimal' if use_udf and nullable else 'Decimal'
        if full_field_type.__contains__('float'):
            full_field_type = full_field_type.replace('float', 'Float')
            simple_field_type = 'UdfFloat'  if use_udf and nullable else 'Float'
        if full_field_type == 'timestamp':
            full_field_type = 'DateTime()'
            simple_field_type = 'DateTime'
        if full_field_type == 'date':
            full_field_type = 'Date()'
            simple_field_type = 'Date'
        if type == 'Model':
            if dbchange :
                db_field_list.append('{field_name} = db.Column(db.{full_field_type}, '
                                 ' comment=\'{comment}\',name=\'{src_field_name}\')'
                                 .format(field_name=field_name, full_field_type=full_field_type,
                                         comment=desc,src_field_name=src_field_name))
            else:
                db_field_list.append('{field_name} = db.Column(db.{full_field_type}, '
                                     ' comment=\'{comment}\')'
                                     .format(field_name=field_name, full_field_type=full_field_type,
                                             comment=desc))

        required = not nullable
        if type == 'Schema':
            fn = 'udf_fields' if use_udf and simple_field_type.__contains__('Udf') else 'fields'
            allow_none = 'allow_none=True, \n' if not required else ''
            v = 'validate=[validate.Length(min=1)], \n' if simple_field_type == 'String' and required else ''
            db_field_list.append("{field_name} = {fn}.{simple_field_type}(\nrequired={required}, \n"
                                 "{allow_none}"
                                 "{v}"
                                 "error_messages=\{\{'required': '{field_desc1}不能为空'\}\}, \n"
                                 "data_key='{src_field_name}')\n"
                                 .format(field_name=field_name,fn=fn, simple_field_type=simple_field_type, required=required,
                                         v=v, allow_none=allow_none,field_desc1=field_desc1, src_field_name=src_field_name))
        if type == 'Factory':
            db_field = ''
            if simple_field_type == 'Decimal':
                db_field = '{field_name} = Decimal(18, 8)'
            if simple_field_type == 'String':
                str_count = int(full_field_type.replace('String(', '').replace(')', ''))
                if str_count == 1:
                    db_field = '{field_name} = \'Y\''
                else:
                    db_field = '{field_name} = factory.Sequence(lambda n: \'{field_name}%d\' % n)'
            elif simple_field_type == 'Integer':
                db_field = '{field_name} = factory.Sequence(lambda n: n)'
            elif simple_field_type == 'DateTime':
                db_field = '{field_name} = datetime.now()'
            db_field_list.append(db_field.format(field_name=field_name))
    for db_field in db_field_list:
        print(db_field)


def print_all_field_info(file_name: None, db_change: False, use_udf: False):
    print_field_info_list('Model',file_name,db_change,use_udf)
    print('---------------------------------------')
    print_field_info_list('Factory',file_name,db_change,use_udf)
    print('---------------------------------------')
    print_field_info_list('Schema',file_name,db_change,use_udf)


#print(camel_to_underline('AbcAss111'))
# print_all_field_info()
# print(re.match('123[(](.*?)123', '123(asdasd123'))
# print(re.match('.*[(][0-9][)]+.*', '(20)'))
# print(re.sub('[(]\\d+[)]',  '', 'int(20)', count=1, ))

def get_label(option): return option.get('label')
if __name__ == '__main__':
    title1 = '选择一个实体excel: '
    options1 = [x for x in os.listdir('.') if os.path.isfile(x) and os.path.splitext(x)[1] == '.xlsx']
    # options = [{'label': 'option1'}, {'label': 'option2'}, {'label': 'option3'}]
    selected1,idx1 = pick(options1, title1, multi_select=False, min_selection_count=1)
    print(selected1)
    title2 = '数据库字段命名类型: '
    options2 = [ {'label': '下划线命名','val': False},{'label': '驼峰','val': True}]
    selected2,idx2 = pick(options2, title2, indicator='*', options_map_func=get_label)
    print(selected2)
    title3 = '是否使用自定义字段类型: '
    options3 = [{'label': '否', 'val': False}, {'label': '是', 'val': True}]
    selected3, idx3 = pick(options3, title3, indicator='*', options_map_func=get_label)
    print(selected3)
    print("\n\n")
    print_all_field_info(file_name=selected1,db_change=selected2.get("val"),use_udf=selected3.get("val"))
